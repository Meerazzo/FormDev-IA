import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convertit un fichier XLSX (colonnes question/reponse) en JSON pour /surveys/forms/analyze."
    )
    parser.add_argument("input_xlsx", help="Chemin vers le fichier XLSX source")
    parser.add_argument(
        "--output",
        default="form_payload.json",
        help="Chemin du fichier JSON de sortie",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Nom de la feuille Excel à utiliser (par défaut : feuille active)",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="Première ligne de données à lire (par défaut : 2, car la ligne 1 contient les en-têtes)",
    )
    parser.add_argument(
        "--end-row",
        type=int,
        default=None,
        help="Dernière ligne de données à lire incluse (par défaut : jusqu'à la dernière ligne non vide)",
    )
    parser.add_argument(
        "--survey-id",
        default="panel_test_001",
        help="Valeur de survey_id dans le JSON généré",
    )
    parser.add_argument(
        "--formation",
        default="panel_reel_test",
        help="Valeur metadata.formation",
    )
    parser.add_argument(
        "--client",
        default="formdev",
        help="Valeur metadata.client",
    )
    args = parser.parse_args()

    input_path = Path(args.input_xlsx)
    output_path = Path(args.output)

    if not input_path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {input_path}")

    wb = load_workbook(filename=input_path, read_only=True, data_only=True)

    if args.sheet:
        if args.sheet not in wb.sheetnames:
            raise ValueError(
                f"Feuille '{args.sheet}' introuvable. Feuilles disponibles : {wb.sheetnames}"
            )
        ws = wb[args.sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    header_question = normalize_text(ws["A1"].value).lower()
    header_reponse = normalize_text(ws["B1"].value).lower()

    if header_question not in {"question", "questions"}:
        raise ValueError(
            f"Colonne A inattendue : '{ws['A1'].value}'. La première cellule doit être 'question'."
        )

    if header_reponse not in {"reponse", "réponse", "reponses", "réponses"}:
        raise ValueError(
            f"Colonne B inattendue : '{ws['B1'].value}'. La première cellule doit être 'reponse' ou 'réponse'."
        )

    start_row = args.start_row
    end_row = args.end_row or ws.max_row

    if start_row < 2:
        raise ValueError("start-row doit être >= 2")
    if end_row < start_row:
        raise ValueError("end-row doit être >= start-row")

    items = []
    item_index = 1

    for row_idx in range(start_row, end_row + 1):
        question_text = normalize_text(ws[f"A{row_idx}"].value)
        response_text = normalize_text(ws[f"B{row_idx}"].value)

        if not question_text and not response_text:
            continue

        items.append(
            {
                "question_id": f"row_{item_index:03d}",
                "question_text": question_text,
                "response_text": response_text,
            }
        )
        item_index += 1

    payload = {
        "survey_id": args.survey_id,
        "items": items,
        "metadata": {
            "formation": args.formation,
            "client": args.client,
            "source_file": input_path.name,
            "sheet": ws.title,
            "start_row": start_row,
            "end_row": end_row,
        },
    }

    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"JSON écrit dans : {output_path}")
    print(f"Nombre d'items exportés : {len(items)}")


if __name__ == "__main__":
    main()